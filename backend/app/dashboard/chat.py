############################################################
#
# mindrouter2 - LLM Inference Translator and Load Balancer
#
# chat.py: Web chat interface routes
#
# Luke Sheneman
# Research Computing and Data Services (RCDS)
# Institute for Interdisciplinary Data Sciences (IIDS)
# University of Idaho
# sheneman@uidaho.edu
#
############################################################

"""Chat interface routes for MindRouter2."""

import asyncio
import base64
import io
import json
import os
import time
from typing import List, Optional, Tuple

from fastapi import APIRouter, Depends, HTTPException, Request, UploadFile, File
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse, RedirectResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from sqlalchemy.ext.asyncio import AsyncSession

from backend.app.core.canonical_schemas import CanonicalModelInfo
from backend.app.core.latex_normalize import normalize_latex
from backend.app.core.telemetry.registry import get_registry
from backend.app.core.translators.openai_in import OpenAIInTranslator
from backend.app.db import crud
from backend.app.db import chat_crud
from backend.app.db.models import ApiKey, User
from backend.app.db.session import get_async_db, get_async_db_context
from backend.app.dashboard.routes import get_session_user_id
from backend.app.logging_config import get_logger
from backend.app.services.inference import InferenceService
from backend.app.settings import get_settings

logger = get_logger(__name__)

chat_router = APIRouter(tags=["chat"])

# Setup templates
templates_path = os.path.join(os.path.dirname(__file__), "templates")
templates = Jinja2Templates(directory=templates_path)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

async def _get_chat_user(
    request: Request, db: AsyncSession
) -> Tuple[User, ApiKey]:
    """Get user and their first active API key from session cookie.

    Returns:
        (User, ApiKey) tuple

    Raises:
        HTTPException 401 if not logged in
        HTTPException 403 if user has no active API key
    """
    user_id = get_session_user_id(request)
    if not user_id:
        raise HTTPException(status_code=401, detail="Not authenticated")

    user = await crud.get_user_by_id(db, user_id)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")

    api_keys = await crud.get_user_api_keys(db, user_id, include_revoked=False)
    if not api_keys:
        raise HTTPException(
            status_code=403,
            detail="No active API key. Create one in your dashboard first.",
        )

    return user, api_keys[0]


def _sharded_path(base_dir: str, attachment_id: int, suffix: str) -> str:
    """Return a sharded filesystem path: base_dir/<id%1000>/<id><suffix>."""
    shard = attachment_id % 1000
    shard_dir = os.path.join(base_dir, str(shard))
    return os.path.join(shard_dir, f"{attachment_id}{suffix}")


def _generate_image_thumbnail(file_bytes: bytes) -> Optional[bytes]:
    """Generate a 200px-wide PNG thumbnail from image bytes.

    Returns raw PNG bytes (not base64).
    """
    from PIL import Image

    img = Image.open(io.BytesIO(file_bytes))
    if img.mode in ("RGBA", "P"):
        img = img.convert("RGB")
    w, h = img.size
    if w > 200:
        new_h = int(h * 200 / w)
        img = img.resize((200, max(new_h, 1)), Image.LANCZOS)
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    return buf.getvalue()


def _generate_medium_thumbnail(file_bytes: bytes) -> bytes:
    """Generate an 800px-wide JPEG medium thumbnail from image bytes.

    Returns raw JPEG bytes.  If the image is already <= 800px wide,
    return a JPEG-compressed copy at the original size (no upscaling).
    """
    from PIL import Image

    img = Image.open(io.BytesIO(file_bytes))
    if img.mode in ("RGBA", "P"):
        img = img.convert("RGB")
    w, h = img.size
    if w > 800:
        new_h = int(h * 800 / w)
        img = img.resize((800, max(new_h, 1)), Image.LANCZOS)
    buf = io.BytesIO()
    img.save(buf, format="JPEG", quality=85)
    return buf.getvalue()


def _generate_pdf_thumbnail(file_bytes: bytes) -> Optional[bytes]:
    """Generate a 200px-wide PNG thumbnail of the first page of a PDF.

    Returns raw PNG bytes (not base64).
    """
    import pdfplumber
    from PIL import Image

    with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
        if not pdf.pages:
            return None
        page = pdf.pages[0]
        page_img = page.to_image(resolution=72)
        pil_img = page_img.original
        w, h = pil_img.size
        if w > 0:
            new_h = int(h * 200 / w)
            pil_img = pil_img.resize((200, max(new_h, 1)), Image.LANCZOS)
        buf = io.BytesIO()
        pil_img.save(buf, format="PNG")
        return buf.getvalue()


def _extract_text_from_docx(file_bytes: bytes) -> str:
    """Extract text from a .docx file."""
    from docx import Document

    doc = Document(io.BytesIO(file_bytes))
    return "\n".join(p.text for p in doc.paragraphs if p.text)


def _extract_text_from_xlsx(file_bytes: bytes) -> str:
    """Extract text from an .xlsx file."""
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    lines = []
    for sheet in wb.worksheets:
        lines.append(f"--- Sheet: {sheet.title} ---")
        for row in sheet.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            lines.append("\t".join(cells))
    wb.close()
    return "\n".join(lines)


def _extract_text_from_pdf(file_bytes: bytes) -> str:
    """Extract text from a PDF file."""
    import pdfplumber

    text_parts = []
    with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
        for i, page in enumerate(pdf.pages):
            page_text = page.extract_text()
            if page_text:
                text_parts.append(f"--- Page {i + 1} ---\n{page_text}")
    return "\n\n".join(text_parts)


def _process_image(file_bytes: bytes) -> bytes:
    """Resize image to max 1536px on longest side and compress as JPEG q85."""
    from PIL import Image

    img = Image.open(io.BytesIO(file_bytes))
    if img.mode in ("RGBA", "P"):
        img = img.convert("RGB")

    max_dim = 1536
    w, h = img.size
    if max(w, h) > max_dim:
        if w >= h:
            new_w = max_dim
            new_h = int(h * max_dim / w)
        else:
            new_h = max_dim
            new_w = int(w * max_dim / h)
        img = img.resize((new_w, max(new_h, 1)), Image.LANCZOS)

    buf = io.BytesIO()
    img.save(buf, format="JPEG", quality=85)
    return buf.getvalue()


def _build_llm_messages(messages_with_attachments, *, model_supports_multimodal: bool = True):
    """Build OpenAI-format messages array from DB messages with attachments.

    Each message dict has: role, content, attachments (list of ChatAttachment ORM objects).
    For images, reads processed JPEG from filesystem and builds image_url blocks.
    For documents, includes extracted_text as context.

    When *model_supports_multimodal* is False, image blocks are replaced with a
    text placeholder so that non-multimodal models never receive image data.
    """
    api_msgs = []
    for msg in messages_with_attachments:
        role = msg.role
        content = msg.content
        attachments = msg.attachments

        if role == "assistant":
            api_msgs.append({"role": "assistant", "content": content or ""})
            continue

        # User message — may have attachments
        if attachments:
            content_blocks = []

            for att in attachments:
                if att.is_image and att.storage_path:
                    if not model_supports_multimodal:
                        content_blocks.append({
                            "type": "text",
                            "text": f"[Image omitted — model does not support multimodal input: {att.filename}]",
                        })
                        continue
                    # Read processed image from filesystem
                    try:
                        with open(att.storage_path, "rb") as f:
                            img_bytes = f.read()
                        b64 = base64.b64encode(img_bytes).decode("utf-8")
                        content_blocks.append({
                            "type": "image_url",
                            "image_url": {
                                "url": f"data:image/jpeg;base64,{b64}",
                            },
                        })
                    except (OSError, IOError):
                        # File missing — skip
                        pass
                elif att.extracted_text:
                    content_blocks.append({
                        "type": "text",
                        "text": f"[File: {att.filename}]\n{att.extracted_text}",
                    })

            if content:
                content_blocks.append({"type": "text", "text": content})

            api_msgs.append({"role": "user", "content": content_blocks})
        else:
            api_msgs.append({"role": "user", "content": content or ""})

    return api_msgs


# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------

@chat_router.get("/chat", response_class=HTMLResponse)
async def chat_page(
    request: Request,
    db: AsyncSession = Depends(get_async_db),
):
    """Serve the chat interface."""
    user_id = get_session_user_id(request)
    if not user_id:
        return RedirectResponse(url="/login", status_code=302)

    user = await crud.get_user_by_id(db, user_id)
    if not user:
        return RedirectResponse(url="/login", status_code=302)

    # Check if user has an API key
    api_keys = await crud.get_user_api_keys(db, user_id, include_revoked=False)
    has_api_key = len(api_keys) > 0

    return templates.TemplateResponse(
        "chat.html",
        {
            "request": request,
            "user": user,
            "has_api_key": has_api_key,
        },
    )


@chat_router.get("/chat/api/models")
async def chat_list_models(
    request: Request,
    db: AsyncSession = Depends(get_async_db),
):
    """List available models for the chat interface (session auth)."""
    user_id = get_session_user_id(request)
    if not user_id:
        return JSONResponse({"error": "Not authenticated"}, status_code=401)

    registry = get_registry()
    backends = await registry.get_healthy_backends()

    model_data: dict = {}
    for backend in backends:
        backend_models = await registry.get_backend_models(backend.id)
        for model in backend_models:
            if model.name not in model_data:
                model_data[model.name] = {
                    "capabilities": {
                        "multimodal": False,
                        "embeddings": False,
                        "structured_output": True,
                        "thinking": False,
                    },
                    "created": int(model.created_at.timestamp()) if model.created_at else int(time.time()),
                }
            if model.supports_multimodal:
                model_data[model.name]["capabilities"]["multimodal"] = True
            if getattr(model, "supports_thinking", False):
                model_data[model.name]["capabilities"]["thinking"] = True
            if "embed" in model.name.lower():
                model_data[model.name]["capabilities"]["embeddings"] = True

    # Read core model config from DB
    core_models = await crud.get_config_json(db, "chat.core_models", [])
    default_model = await crud.get_config_json(db, "chat.default_model", None)

    models = []
    for name, data in sorted(model_data.items()):
        # Skip embedding-only models
        if data["capabilities"]["embeddings"] and not data["capabilities"]["multimodal"]:
            continue
        models.append({
            "id": name,
            "capabilities": data["capabilities"],
            "is_core": name in core_models,
        })

    # Check whether any core model is currently online
    online_names = {m["id"] for m in models}
    has_core = any(cm in online_names for cm in core_models)

    return JSONResponse({
        "models": models,
        "default_model": default_model,
        "has_core": has_core,
    })


# ---------------------------------------------------------------------------
# Conversation CRUD endpoints
# ---------------------------------------------------------------------------

@chat_router.get("/chat/api/conversations")
async def list_conversations(
    request: Request,
    db: AsyncSession = Depends(get_async_db),
):
    """List user's conversations."""
    user_id = get_session_user_id(request)
    if not user_id:
        return JSONResponse({"error": "Not authenticated"}, status_code=401)

    convs = await chat_crud.get_user_conversations(db, user_id)
    return JSONResponse({
        "conversations": [
            {
                "id": c.id,
                "title": c.title,
                "model": c.model,
                "updated_at": c.updated_at.isoformat() if c.updated_at else None,
            }
            for c in convs
        ]
    })


@chat_router.post("/chat/api/conversations")
async def create_conversation(
    request: Request,
    db: AsyncSession = Depends(get_async_db),
):
    """Create a new conversation."""
    user_id = get_session_user_id(request)
    if not user_id:
        return JSONResponse({"error": "Not authenticated"}, status_code=401)

    try:
        body = await request.json()
    except Exception:
        body = {}

    model = body.get("model")
    conv = await chat_crud.create_conversation(db, user_id, model=model)
    await db.commit()

    return JSONResponse({
        "id": conv.id,
        "title": conv.title,
        "model": conv.model,
    })


@chat_router.get("/chat/api/conversations/{conversation_id}")
async def get_conversation(
    conversation_id: int,
    request: Request,
    db: AsyncSession = Depends(get_async_db),
):
    """Get a conversation with all messages."""
    user_id = get_session_user_id(request)
    if not user_id:
        return JSONResponse({"error": "Not authenticated"}, status_code=401)

    data = await chat_crud.get_conversation_with_messages(db, conversation_id, user_id)
    if not data:
        return JSONResponse({"error": "Conversation not found"}, status_code=404)

    return JSONResponse(data)


@chat_router.delete("/chat/api/conversations/{conversation_id}")
async def delete_conversation(
    conversation_id: int,
    request: Request,
    db: AsyncSession = Depends(get_async_db),
):
    """Delete a conversation and its messages/attachments."""
    user_id = get_session_user_id(request)
    if not user_id:
        return JSONResponse({"error": "Not authenticated"}, status_code=401)

    ok = await chat_crud.delete_conversation(db, conversation_id, user_id)
    if not ok:
        return JSONResponse({"error": "Conversation not found"}, status_code=404)

    await db.commit()
    return JSONResponse({"ok": True})


@chat_router.patch("/chat/api/conversations/{conversation_id}")
async def patch_conversation(
    conversation_id: int,
    request: Request,
    db: AsyncSession = Depends(get_async_db),
):
    """Update conversation title or model."""
    user_id = get_session_user_id(request)
    if not user_id:
        return JSONResponse({"error": "Not authenticated"}, status_code=401)

    try:
        body = await request.json()
    except Exception:
        return JSONResponse({"error": "Invalid JSON"}, status_code=400)

    conv = await chat_crud.update_conversation(
        db, conversation_id, user_id,
        title=body.get("title"),
        model=body.get("model"),
    )
    if not conv:
        return JSONResponse({"error": "Conversation not found"}, status_code=404)

    await db.commit()
    return JSONResponse({
        "id": conv.id,
        "title": conv.title,
        "model": conv.model,
    })


# ---------------------------------------------------------------------------
# Upload (modified: stores to DB + filesystem)
# ---------------------------------------------------------------------------

@chat_router.post("/chat/api/upload")
async def chat_upload(
    request: Request,
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_async_db),
):
    """Upload a file, process it, and store to DB + filesystem."""
    user_id = get_session_user_id(request)
    if not user_id:
        return JSONResponse({"error": "Not authenticated"}, status_code=401)

    settings = get_settings()
    max_size = settings.chat_upload_max_size_mb * 1024 * 1024

    # Read file content
    file_bytes = await file.read()
    if len(file_bytes) > max_size:
        return JSONResponse(
            {"error": f"File too large. Maximum size is {settings.chat_upload_max_size_mb}MB."},
            status_code=413,
        )

    filename = file.filename or "unknown"
    ext = os.path.splitext(filename)[1].lower()

    # Check allowed extensions
    allowed = settings.chat_upload_allowed_extensions
    if ext not in allowed:
        return JSONResponse(
            {"error": f"File type '{ext}' not supported. Allowed: {', '.join(allowed)}"},
            status_code=400,
        )

    image_extensions = {".jpg", ".jpeg", ".png", ".gif", ".webp"}
    is_image = ext in image_extensions
    storage_path = None
    thumbnail = None
    extracted_text = None
    content_type = file.content_type or ("image/jpeg" if is_image else "application/octet-stream")

    if is_image:
        # Process image: resize + compress to JPEG
        try:
            processed = await asyncio.to_thread(_process_image, file_bytes)
        except Exception:
            processed = file_bytes

        # Generate thumbnail (raw PNG bytes)
        thumb_bytes = None
        try:
            thumb_bytes = await asyncio.to_thread(_generate_image_thumbnail, processed)
        except Exception:
            pass

        # Create attachment record first to get ID for filename
        att = await chat_crud.create_attachment(
            db, user_id,
            filename=filename,
            content_type="image/jpeg",
            is_image=True,
            file_size=len(processed),
        )

        # Store processed image to filesystem (sharded directory)
        files_dir = settings.chat_files_path
        storage_path = _sharded_path(files_dir, att.id, ".jpg")
        os.makedirs(os.path.dirname(storage_path), exist_ok=True)
        await asyncio.to_thread(_write_file, storage_path, processed)

        # Save thumbnail to filesystem
        thumb_url = None
        if thumb_bytes:
            thumb_path = _sharded_path(files_dir, att.id, "_thumb.png")
            await asyncio.to_thread(_write_file, thumb_path, thumb_bytes)
            att.thumbnail_path = thumb_path
            thumb_url = f"/chat/api/attachments/{att.id}/thumbnail"

        # Generate and save medium thumbnail (800px wide) for modal display
        try:
            medium_bytes = await asyncio.to_thread(_generate_medium_thumbnail, processed)
            medium_path = _sharded_path(files_dir, att.id, "_medium.jpg")
            await asyncio.to_thread(_write_file, medium_path, medium_bytes)
        except Exception:
            pass  # Non-critical — modal will fall back to full image

        att.storage_path = storage_path
        await db.flush()
        await db.commit()

        return JSONResponse({
            "attachment_id": att.id,
            "filename": filename,
            "content_type": "image/jpeg",
            "is_image": True,
            "thumbnail": thumb_url,
        })

    # Text extraction for non-image files
    try:
        if ext in (".txt", ".md", ".csv", ".log"):
            extracted_text = file_bytes.decode("utf-8", errors="replace")
        elif ext == ".json":
            extracted_text = file_bytes.decode("utf-8", errors="replace")
        elif ext in (".html", ".htm"):
            extracted_text = file_bytes.decode("utf-8", errors="replace")
        elif ext == ".docx":
            extracted_text = await asyncio.to_thread(_extract_text_from_docx, file_bytes)
        elif ext == ".xlsx":
            extracted_text = await asyncio.to_thread(_extract_text_from_xlsx, file_bytes)
        elif ext == ".pdf":
            extracted_text = await asyncio.to_thread(_extract_text_from_pdf, file_bytes)
        else:
            extracted_text = file_bytes.decode("utf-8", errors="replace")
    except Exception as e:
        logger.warning("file_extraction_error", filename=filename, error=str(e))
        return JSONResponse(
            {"error": f"Failed to extract text from '{filename}': {e}"},
            status_code=422,
        )

    # Generate thumbnail for PDFs (raw bytes)
    thumb_bytes = None
    if ext == ".pdf":
        try:
            thumb_bytes = await asyncio.to_thread(_generate_pdf_thumbnail, file_bytes)
        except Exception:
            pass

    att = await chat_crud.create_attachment(
        db, user_id,
        filename=filename,
        content_type=content_type,
        is_image=False,
        extracted_text=extracted_text,
        file_size=len(file_bytes),
    )

    # Save PDF thumbnail to filesystem
    thumb_url = None
    if thumb_bytes:
        files_dir = settings.chat_files_path
        thumb_path = _sharded_path(files_dir, att.id, "_thumb.png")
        os.makedirs(os.path.dirname(thumb_path), exist_ok=True)
        await asyncio.to_thread(_write_file, thumb_path, thumb_bytes)
        att.thumbnail_path = thumb_path
        thumb_url = f"/chat/api/attachments/{att.id}/thumbnail"
        await db.flush()

    await db.commit()

    return JSONResponse({
        "attachment_id": att.id,
        "filename": filename,
        "content_type": content_type,
        "is_image": False,
        "thumbnail": thumb_url,
    })


def _write_file(path: str, data: bytes):
    """Write bytes to a file (used in asyncio.to_thread)."""
    with open(path, "wb") as f:
        f.write(data)


# ---------------------------------------------------------------------------
# Medium thumbnail serving
# ---------------------------------------------------------------------------

@chat_router.get("/chat/api/attachments/{attachment_id}/medium")
async def serve_medium_thumbnail(
    attachment_id: int,
    request: Request,
    db: AsyncSession = Depends(get_async_db),
):
    """Serve the 800px medium thumbnail for an image attachment."""
    user_id = get_session_user_id(request)
    if not user_id:
        raise HTTPException(status_code=401, detail="Not authenticated")

    att = await chat_crud.get_attachment(db, attachment_id, user_id)
    if not att or not att.is_image:
        raise HTTPException(status_code=404, detail="Attachment not found")

    settings = get_settings()

    # Try sharded path first, then legacy flat path, then full image fallback
    medium_path = _sharded_path(settings.chat_files_path, att.id, "_medium.jpg")
    if not os.path.exists(medium_path):
        legacy_medium = os.path.join(settings.chat_files_path, f"{att.id}_medium.jpg")
        if os.path.exists(legacy_medium):
            medium_path = legacy_medium
        elif att.storage_path and os.path.exists(att.storage_path):
            medium_path = att.storage_path
        else:
            raise HTTPException(status_code=404, detail="Image file not found")

    return FileResponse(medium_path, media_type="image/jpeg")


# ---------------------------------------------------------------------------
# Thumbnail serving
# ---------------------------------------------------------------------------

@chat_router.get("/chat/api/attachments/{attachment_id}/thumbnail")
async def serve_thumbnail(
    attachment_id: int,
    request: Request,
    db: AsyncSession = Depends(get_async_db),
):
    """Serve the small PNG thumbnail for an attachment."""
    user_id = get_session_user_id(request)
    if not user_id:
        raise HTTPException(status_code=401, detail="Not authenticated")

    att = await chat_crud.get_attachment(db, attachment_id, user_id)
    if not att:
        raise HTTPException(status_code=404, detail="Attachment not found")

    # Serve filesystem thumbnail
    if att.thumbnail_path and os.path.exists(att.thumbnail_path):
        return FileResponse(att.thumbnail_path, media_type="image/png")

    raise HTTPException(status_code=404, detail="Thumbnail not found")


# ---------------------------------------------------------------------------
# Completions (modified: server-side message building from DB)
# ---------------------------------------------------------------------------

async def _save_assistant_message(conversation_id: int, content: str):
    """Save an assistant message using an independent DB session.

    Called from the streaming generator's finally block via asyncio.shield()
    so it survives ASGI task cancellation.
    """
    async with get_async_db_context() as save_db:
        await chat_crud.create_message(save_db, conversation_id, "assistant", content)
        await save_db.commit()


@chat_router.post("/chat/api/completions")
async def chat_completions(
    request: Request,
    db: AsyncSession = Depends(get_async_db),
):
    """Chat completions with server-side conversation management."""
    try:
        user, api_key = await _get_chat_user(request, db)
    except HTTPException as e:
        return JSONResponse({"error": e.detail}, status_code=e.status_code)

    try:
        body = await request.json()
    except Exception:
        return JSONResponse({"error": "Invalid JSON body"}, status_code=400)

    conversation_id = body.get("conversation_id")
    content = body.get("content", "").strip()
    attachment_ids = body.get("attachment_ids", [])
    model = body.get("model")
    stream = body.get("stream", True)

    if not conversation_id:
        return JSONResponse({"error": "conversation_id is required"}, status_code=400)
    if not content and not attachment_ids:
        return JSONResponse({"error": "content or attachment_ids required"}, status_code=400)
    if not model:
        return JSONResponse({"error": "model is required"}, status_code=400)

    # Verify conversation ownership
    conv = await chat_crud.get_conversation(db, conversation_id, user.id)
    if not conv:
        return JSONResponse({"error": "Conversation not found"}, status_code=404)

    # Update conversation model
    conv.model = model
    await db.flush()

    # Create user message
    user_msg = await chat_crud.create_message(db, conversation_id, "user", content or None)

    # Link attachments to the user message
    if attachment_ids:
        await chat_crud.link_attachments_to_message(db, attachment_ids, user_msg.id, user.id)

    # Auto-title conversation from first user message
    if conv.title == "New Chat" and content:
        title = content[:60] + ("..." if len(content) > 60 else "")
        conv.title = title
        await db.flush()

    # Commit user message + attachment links before streaming
    await db.commit()

    # Load full conversation history from DB
    messages = await chat_crud.get_conversation_messages(db, conversation_id)

    # Look up whether the target model supports vision
    model_supports_multimodal = False
    registry = get_registry()
    healthy_backends = await registry.get_healthy_backends()
    for b in healthy_backends:
        b_models = await registry.get_backend_models(b.id)
        for m in b_models:
            if m.name == model and m.supports_multimodal:
                model_supports_multimodal = True
                break
        if model_supports_multimodal:
            break

    # Build OpenAI messages array from DB records
    api_messages = _build_llm_messages(messages, model_supports_multimodal=model_supports_multimodal)

    # Inject system prompt for consistent math formatting
    # (prepend so it appears before conversation history)
    _SYSTEM_PROMPT = (
        "You are a helpful AI assistant provided by the University of Idaho. "
        "All conversations are processed entirely on University of Idaho infrastructure — "
        "your messages never leave campus servers and are not shared with any third party. "
        "If users ask about data privacy, reassure them of this.\n\n"
        "When writing mathematical expressions, always use LaTeX with proper "
        "delimiters. Use $...$ for inline math and $$...$$ for display equations. "
        "Never leave LaTeX commands like \\frac, \\int, \\sum, \\alpha, etc. "
        "bare without dollar-sign delimiters. Always wrap complete expressions, "
        "not individual symbols — write $\\frac{a}{b} + c$ not $\\frac{a}{b}$ + c."
    )
    # Only add if there is no user-supplied system message already
    has_system = any(m.get("role") == "system" for m in api_messages)
    if not has_system:
        api_messages.insert(0, {"role": "system", "content": _SYSTEM_PROMPT})
    else:
        # Append math instructions to the existing system message
        for m in api_messages:
            if m.get("role") == "system":
                m["content"] = m["content"] + "\n\n" + _SYSTEM_PROMPT
                break

    # Build canonical request
    try:
        request_data = {
            "model": model,
            "messages": api_messages,
            "stream": stream,
            "max_tokens": 16384,
        }
        # Pass thinking parameters if provided
        think = body.get("think")
        reasoning_effort = body.get("reasoning_effort")
        if think is not None:
            request_data["think"] = think
        if reasoning_effort is not None:
            request_data["reasoning_effort"] = reasoning_effort

        canonical = OpenAIInTranslator.translate_chat_request(request_data)
    except Exception as e:
        return JSONResponse({"error": f"Invalid request: {e}"}, status_code=400)

    service = InferenceService(db)

    if canonical.stream:
        # We need to capture the full response to save as assistant message
        async def generate():
            full_content = ""
            try:
                async for chunk in service.stream_chat_completion(
                    canonical, user, api_key, request
                ):
                    # Parse chunk to capture content
                    if isinstance(chunk, bytes):
                        chunk_str = chunk.decode("utf-8", errors="replace")
                    else:
                        chunk_str = chunk

                    # Extract content from SSE data for saving
                    for line in chunk_str.split("\n"):
                        line = line.strip()
                        if line.startswith("data: ") and line != "data: [DONE]":
                            try:
                                import json
                                parsed = json.loads(line[6:])
                                delta = parsed.get("choices", [{}])[0].get("delta", {}).get("content", "")
                                if delta:
                                    full_content += delta
                            except (json.JSONDecodeError, IndexError, KeyError):
                                pass

                    yield chunk
            except HTTPException as e:
                detail = e.detail if isinstance(e.detail, str) else json.dumps(e.detail)
                error_data = "data: " + json.dumps({"error": detail}) + "\n\n"
                yield error_data.encode()
                yield b"data: [DONE]\n\n"
            except Exception as e:
                logger.exception("chat_stream_error", error=str(e))
                error_data = "data: " + json.dumps({"error": "Internal server error"}) + "\n\n"
                yield error_data.encode()
                yield b"data: [DONE]\n\n"
            finally:
                # Save assistant message using an independent DB session,
                # shielded from ASGI cancellation.  When the client finishes
                # reading the stream, uvicorn may cancel the task — we must
                # protect the DB write from that.
                if full_content:
                    try:
                        await asyncio.shield(_save_assistant_message(
                            conversation_id, normalize_latex(full_content)
                        ))
                    except asyncio.CancelledError:
                        pass
                    except Exception as e:
                        logger.warning("failed_to_save_assistant_message", error=str(e))

        return StreamingResponse(
            generate(),
            media_type="text/event-stream",
            headers={
                "Cache-Control": "no-cache",
                "Connection": "keep-alive",
                "X-Accel-Buffering": "no",
            },
        )
    else:
        try:
            result = await service.chat_completion(
                canonical, user, api_key, request
            )
            # Save assistant message
            assistant_content = ""
            try:
                choices = result.get("choices", [])
                if choices:
                    assistant_content = choices[0].get("message", {}).get("content", "")
            except (KeyError, IndexError):
                pass

            if assistant_content:
                # Normalize LaTeX before saving and returning
                assistant_content = normalize_latex(assistant_content)
                result["choices"][0]["message"]["content"] = assistant_content
                await chat_crud.create_message(db, conversation_id, "assistant", assistant_content)
                await db.commit()

            return JSONResponse(result)
        except HTTPException as e:
            detail = e.detail if isinstance(e.detail, str) else json.dumps(e.detail)
            return JSONResponse({"error": detail}, status_code=e.status_code)
        except Exception as e:
            logger.exception("chat_completion_error", error=str(e))
            return JSONResponse({"error": "Internal server error"}, status_code=500)
